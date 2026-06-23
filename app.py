import streamlit as st
import requests
from bs4 import BeautifulSoup
from openpyxl import load_workbook
from concurrent.futures import ThreadPoolExecutor, as_completed
from openpyxl.styles import PatternFill
import time

st.title("URL更新チェック")

# KW入力
col1, col2 = st.columns(2)

with col1:
    kw1_input = st.text_area("KW①※改行でOR条件")

with col2:
    kw2_input = st.text_area("KW②※改行でOR条件")

uploaded_file = st.file_uploader("Excelアップロード", type=["xlsx"])
st.caption("G列にURL / B列に結果を書き込みします。")

start = st.button("チェック開始")

# ✅ 色設定
pink_fill = PatternFill(start_color="FFC0CB", end_color="FFC0CB", fill_type="solid")
blue_fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")
green_fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")

# ✅ Session
session = requests.Session()
session.headers.update({"User-Agent": "Mozilla/5.0"})

# ✅ リトライ
def fetch_with_retry(url, retries=3):
    for attempt in range(retries):
        try:
            res = session.get(url, timeout=10)
            res.raise_for_status()
            return res
        except:
            if attempt < retries - 1:
                time.sleep(0.5)
            else:
                return None

def check_url(row_idx, url, kw1_list, kw2_list):
    try:
        if not url:
            return (row_idx, "URLなし", "none")

        url = str(url).strip()
        if not url.startswith("http"):
            url = "https://" + url

        time.sleep(0.1)

        response = fetch_with_retry(url)

        if response is None:
            return (row_idx, "取得失敗", "fail")

        soup = BeautifulSoup(response.text, "html.parser")
        page_text = soup.get_text().lower()

        hit1 = [kw for kw in kw1_list if kw.lower() in page_text]
        hit2 = [kw for kw in kw2_list if kw.lower() in page_text]

        # ✅ 判定ロジック
        if hit1 and hit2:
            return (row_idx, "KW①,KW②", "both")
        elif hit1:
            return (row_idx, "KW①", "kw1")
        elif hit2:
            return (row_idx, "KW②", "kw2")
        else:
            return (row_idx, "該当なし", "none")

    except:
        return (row_idx, "取得失敗", "fail")


# ===== 実行 =====
if uploaded_file and start:

    kw1_list = [kw.strip() for kw in kw1_input.split("\n") if kw.strip()]
    kw2_list = [kw.strip() for kw in kw2_input.split("\n") if kw.strip()]

    st.write(f"KW①：{len(kw1_list)}件 / KW②：{len(kw2_list)}件")

    wb = load_workbook(uploaded_file)
    ws = wb.active

    max_row = ws.max_row

    progress = st.progress(0)

    results = {}

    max_workers = 5

    with ThreadPoolExecutor(max_workers=max_workers) as executor:
        futures = [
            executor.submit(check_url, i, ws[f"G{i}"].value, kw1_list, kw2_list)
            for i in range(2, max_row + 1)
        ]

        for i, future in enumerate(as_completed(futures)):
            row_idx, result, status = future.result()
            results[row_idx] = (result, status)
            progress.progress((i + 1) / len(futures))

    # ✅ 書き込み＋色付け
    for row_idx, (result, status) in results.items():
        cell = ws[f"B{row_idx}"]
        cell.value = str(result)

        if status == "kw1":
            cell.fill = pink_fill
        elif status == "kw2":
            cell.fill = blue_fill
        elif status == "both":
            cell.fill = green_fill
        else:
            cell.fill = PatternFill()  # 無色

    st.success("チェック完了🚀")

    # ✅ 表示
    display_data = []
    hit1_count = 0
    hit2_count = 0
    both_count = 0
    fail_count = 0

    for i in range(2, max_row + 1):
        url = ws[f"G{i}"].value
        result = str(ws[f"B{i}"].value)

        if result == "取得失敗":
            fail_count += 1
        elif result == "KW①":
            hit1_count += 1
        elif result == "KW②":
            hit2_count += 1
        elif result == "KW①,KW②":
            both_count += 1

        display_data.append({
            "行": i,
            "URL": url,
            "判定結果": result
        })

    st.subheader("チェック結果")
    st.dataframe(display_data)

    st.write(f"🌸 KW①のみ：{hit1_count}件")
    st.write(f"🔵 KW②のみ：{hit2_count}件")
    st.write(f"🟢 両方：{both_count}件")
    st.write(f"❌ 取得失敗：{fail_count}件")
    st.write(f"📊 全体：{len(display_data)}件")

    # ✅ DL
    output_file = "result.xlsx"
    wb.save(output_file)

    with open(output_file, "rb") as f:
        st.download_button(
            "Excelダウンロード",
            f,
            file_name=output_file
        )
